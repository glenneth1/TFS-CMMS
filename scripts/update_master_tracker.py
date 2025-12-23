#!/usr/bin/env python3
"""
Update master_deficiencies table from Master.xlsx spreadsheet.
"""

import sqlite3
import openpyxl
from datetime import datetime
import sys

DB_PATH = 'data/tfs-cmms.db'
XLSX_PATH = 'reference/Master.xlsx'

def get_camp_mapping(conn):
    """Get camp name to ID mapping."""
    cursor = conn.execute("SELECT id, name FROM camps")
    # Create case-insensitive mapping
    mapping = {}
    for row in cursor:
        mapping[row[1].lower()] = row[0]
        mapping[row[1]] = row[0]
    return mapping

def normalize_camp_name(name):
    """Normalize camp name for matching."""
    if not name:
        return None
    name = str(name).strip()
    # Handle common variations
    name_lower = name.lower()
    if 'asad' in name_lower:
        return 'Al Asad'
    if 'erbil' in name_lower:
        return 'Erbil'
    if 'arifjan' in name_lower:
        return 'Arifjan'
    if 'buehring' in name_lower:
        return 'Buehring'
    return name

def format_date(val):
    """Format date value to YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    return str(val)

def main():
    print(f"Loading {XLSX_PATH}...")
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb['Data']
    
    conn = sqlite3.connect(DB_PATH)
    camp_mapping = get_camp_mapping(conn)
    
    # Get existing deficiency numbers for comparison
    cursor = conn.execute("SELECT deficiency_number FROM master_deficiencies")
    existing = set(row[0] for row in cursor if row[0])
    print(f"Existing records: {len(existing)}")
    
    # Track stats
    new_count = 0
    updated_count = 0
    skipped_count = 0
    unknown_camps = set()
    
    # Process rows (skip header)
    total_rows = ws.max_row - 1
    print(f"Processing {total_rows} rows from spreadsheet...")
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row_num % 10000 == 0:
            print(f"  Processed {row_num - 1} rows...")
        
        camp_name = row[0]  # CAMP
        building_number = row[1]  # Building Number
        deficiency_number = row[2]  # Deficiency Number
        def_category = row[3]  # DEF Category
        equipment = row[4]  # Equipment
        inspection_phase = row[5]  # Inspection phase
        repair_by = row[6]  # Repair By
        occurrences = row[7] if row[7] else 0  # Occurances
        lhs_imminent = row[8]  # LHS immenent
        om_sor_number = row[9]  # O&M SOR #
        repair_team_number = row[10]  # Repair Team #
        parts_ordered = row[11]  # Parts Ordered
        date_ordered = format_date(row[12])  # Date Ordered
        deficiency_status = row[13]  # Deficiency Status
        inspection_date = format_date(row[14])  # Date
        rac = row[15]  # RAC
        
        if not deficiency_number:
            skipped_count += 1
            continue
        
        # Get camp ID
        normalized_camp = normalize_camp_name(camp_name)
        camp_id = camp_mapping.get(normalized_camp) or camp_mapping.get(normalized_camp.lower() if normalized_camp else '')
        
        if not camp_id and camp_name:
            unknown_camps.add(camp_name)
            skipped_count += 1
            continue
        
        if not camp_id:
            skipped_count += 1
            continue
        
        if deficiency_number in existing:
            # Update existing record
            conn.execute("""
                UPDATE master_deficiencies SET
                    building_number = ?,
                    def_category = ?,
                    equipment = ?,
                    inspection_phase = ?,
                    repair_by = ?,
                    occurrences = ?,
                    lhs_imminent = ?,
                    om_sor_number = ?,
                    repair_team_number = ?,
                    parts_ordered = ?,
                    date_ordered = ?,
                    deficiency_status = ?,
                    inspection_date = ?,
                    rac = ?,
                    updated_at = datetime('now')
                WHERE deficiency_number = ?
            """, (building_number, def_category, equipment, inspection_phase,
                  repair_by, occurrences, lhs_imminent, om_sor_number,
                  repair_team_number, parts_ordered, date_ordered,
                  deficiency_status, inspection_date, rac, deficiency_number))
            updated_count += 1
        else:
            # Insert new record
            conn.execute("""
                INSERT INTO master_deficiencies 
                (camp_id, building_number, deficiency_number, def_category, equipment,
                 inspection_phase, repair_by, occurrences, lhs_imminent, om_sor_number,
                 repair_team_number, parts_ordered, date_ordered, deficiency_status,
                 inspection_date, rac)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (camp_id, building_number, deficiency_number, def_category, equipment,
                  inspection_phase, repair_by, occurrences, lhs_imminent, om_sor_number,
                  repair_team_number, parts_ordered, date_ordered, deficiency_status,
                  inspection_date, rac))
            new_count += 1
    
    conn.commit()
    conn.close()
    wb.close()
    
    print(f"\nDone!")
    print(f"  New records: {new_count}")
    print(f"  Updated records: {updated_count}")
    print(f"  Skipped: {skipped_count}")
    
    if unknown_camps:
        print(f"\nUnknown camps (not in database):")
        for camp in sorted(unknown_camps):
            print(f"  - {camp}")

if __name__ == '__main__':
    main()
