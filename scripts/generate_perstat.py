#!/usr/bin/env python3
"""
Generate PERSTAT reports in PDF and Excel formats.
Matches the exact format of:
- Internal_PERSTAT-DSR_TFS_VERSAR (PDF)
- USACE OCONUS PERSTAT (Excel)

Usage:
    python generate_perstat.py --format pdf --output /path/to/output.pdf
    python generate_perstat.py --format excel --output /path/to/output.xlsx
"""

import argparse
import sqlite3
import os
from datetime import datetime
from pathlib import Path

# Get the database path
SCRIPT_DIR = Path(__file__).parent
DB_PATH = SCRIPT_DIR.parent / "data" / "tfs-cmms.db"


def get_personnel_data():
    """Fetch all active personnel from database with user details."""
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            p.*,
            c.name as camp_name,
            co.name as country_name,
            u.team_number,
            u.electrician_type,
            u.hire_date,
            u.bog_date,
            u.role as user_role
        FROM personnel p
        LEFT JOIN camps c ON p.current_camp_id = c.id
        LEFT JOIN countries co ON c.country_id = co.id
        LEFT JOIN users u ON p.user_id = u.id
        WHERE p.status = 'Active'
        ORDER BY p.full_name
    """)
    
    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]


def split_name(full_name):
    """Split full name into first and last name."""
    parts = full_name.strip().split(' ', 1)
    if len(parts) == 2:
        return parts[1], parts[0]  # Last, First
    return full_name, ''


def get_position(role_category, electrician_type):
    """Map role category to position name."""
    if role_category == 'Electrician':
        return 'ELECTRICIAN'
    elif role_category == 'QC':
        return 'QC'
    elif role_category == 'PMO':
        return 'PMO'
    elif role_category == 'Property/Materials':
        return 'Property'
    return role_category or ''


def format_date(date_str):
    """Format date string for display."""
    if not date_str:
        return ''
    try:
        if isinstance(date_str, str):
            # Try parsing various formats
            for fmt in ['%Y-%m-%d', '%Y-%m-%d %H:%M:%S', '%d-%b-%y']:
                try:
                    dt = datetime.strptime(date_str.split()[0], fmt.split()[0])
                    return dt.strftime('%-d-%b-%y')
                except:
                    continue
        return str(date_str)
    except:
        return str(date_str) if date_str else ''


def get_status_color(status):
    """Get background color based on personnel status."""
    status_lower = (status or '').lower()
    if 'r/r' in status_lower or 'r&r' in status_lower:
        return colors.HexColor('#fff3cd')  # Yellow - R&R
    elif 'remote' in status_lower:
        return colors.HexColor('#d1ecf1')  # Light blue - Remote
    elif 'lwop' in status_lower or 'leave without pay' in status_lower:
        return colors.HexColor('#f8d7da')  # Light red - LWOP
    elif 'medical' in status_lower:
        return colors.HexColor('#f8d7da')  # Light red - Medical
    elif 'leave' in status_lower:
        return colors.HexColor('#fff3cd')  # Yellow - Leave
    return None  # No special color (Present)


def generate_pdf_report(output_path):
    """Generate Internal PERSTAT-DSR PDF report matching the example format."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    
    personnel = get_personnel_data()
    report_date = datetime.now().strftime("%d-%b-%y")
    
    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(letter),
        rightMargin=0.3*inch,
        leftMargin=0.3*inch,
        topMargin=0.4*inch,
        bottomMargin=0.4*inch
    )
    
    styles = getSampleStyleSheet()
    elements = []
    
    # TF SAFE CMMS Header with logos
    header_data = [['', 'TF SAFE CMMS', '', 'REPORT DATE:', report_date]]
    
    # Try to add logos if they exist
    logo_path = SCRIPT_DIR.parent / "static" / "img" / "TFS_Logo.png"
    company_logo_path = SCRIPT_DIR.parent / "static" / "img" / "company_logo.png"
    
    if logo_path.exists():
        try:
            logo = Image(str(logo_path), width=0.6*inch, height=0.6*inch)
            header_data[0][0] = logo
        except:
            header_data[0][0] = ''
    
    if company_logo_path.exists():
        try:
            company_logo = Image(str(company_logo_path), width=0.8*inch, height=0.4*inch)
            header_data[0][2] = company_logo
        except:
            header_data[0][2] = ''
    
    header_table = Table(header_data, colWidths=[0.8*inch, 3*inch, 1*inch, 4*inch, 1*inch])
    header_table.setStyle(TableStyle([
        ('FONTNAME', (1, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (1, 0), (1, 0), 16),
        ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ('ALIGN', (3, 0), (4, 0), 'RIGHT'),
        ('FONTSIZE', (3, 0), (4, 0), 10),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 5))
    
    # Subtitle
    subtitle = Paragraph("DAILY STATUS REPORT (PERSTAT)", 
                        ParagraphStyle('Subtitle', fontSize=12, alignment=TA_CENTER))
    elements.append(subtitle)
    elements.append(Spacer(1, 10))
    
    # Headers matching the exact PDF format
    headers = [
        'LAST NAME', 'FIRST NAME', 'ASSIGNED\nLOCATION', 'STATUS', 
        'PHYSICAL LOCATION', 'POSITION', 'NATIONALITY', 
        'ME', 'JE', 'TEAM', 'DEPLOY START\nDATE', 'DEPART DATE'
    ]
    
    # Build table data
    table_data = [headers]
    
    for p in personnel:
        last_name, first_name = split_name(p.get('full_name', ''))
        
        # Determine ME/JE based on electrician type
        elec_type = p.get('electrician_type', '')
        me = 'ME' if elec_type == 'Master' else ''
        je = 'JE' if elec_type == 'Journeyman' else ''
        
        # Get location - use country or current_location
        assigned_loc = p.get('country_name') or p.get('current_location') or 'TBD'
        
        # Physical location - camp name or detailed location
        physical_loc = p.get('camp_name') or p.get('current_location') or ''
        
        # Position
        position = get_position(p.get('role_category'), elec_type)
        
        # Nationality - default based on common patterns
        nationality = p.get('nationality') or 'US'
        
        # Status - default to Present
        status = p.get('status', 'Present')
        if status == 'Active':
            status = 'Present'
        
        # Team number
        team = str(p.get('team_number') or '') if p.get('team_number') else ''
        
        # Dates
        deploy_date = format_date(p.get('bog_date'))
        depart_date = ''  # Usually empty unless on R&R
        
        row = [
            last_name,
            first_name,
            assigned_loc,
            status,
            physical_loc,
            position,
            nationality,
            me,
            je,
            team,
            deploy_date,
            depart_date
        ]
        table_data.append(row)
    
    # Create table with column widths matching the PDF example
    col_widths = [1.1*inch, 0.9*inch, 0.65*inch, 0.55*inch, 1.4*inch, 
                  0.75*inch, 0.6*inch, 0.25*inch, 0.25*inch, 0.4*inch, 0.75*inch, 0.7*inch]
    
    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    
    # Build style list
    style_commands = [
        # Header styling - dark blue background
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a365d')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 6),
        
        # Data rows
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 1), (1, -1), 'LEFT'),  # Names left aligned
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),  # Rest centered
        ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
        ('TOPPADDING', (0, 1), (-1, -1), 4),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
    ]
    
    # Apply conditional row coloring based on status (column 3, index 3)
    for row_idx, row in enumerate(table_data[1:], start=1):  # Skip header
        status = row[3] if len(row) > 3 else ''
        status_color = get_status_color(status)
        if status_color:
            style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), status_color))
        else:
            # Alternating white/light gray for Present rows
            bg_color = colors.white if row_idx % 2 == 1 else colors.HexColor('#f5f5f5')
            style_commands.append(('BACKGROUND', (0, row_idx), (-1, row_idx), bg_color))
    
    table.setStyle(TableStyle(style_commands))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    print(f"PDF report generated: {output_path}")


def generate_excel_report(output_path):
    """Generate USACE OCONUS PERSTAT Excel report matching the DSR format."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    
    personnel = get_personnel_data()
    report_date = datetime.now().strftime("%Y-%m-%d")
    report_date_display = datetime.now().strftime("%d-%b-%y")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "DSR"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center')
    
    # Title section (matching USACE format)
    ws['A1'] = "DAILY STATUS REPORT"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['B2'] = "REPORTING UNIT:"
    ws['C2'] = "Huntsville Center - TF SAFE"
    
    ws['B3'] = "REPORT DATE:"
    ws['C3'] = report_date_display
    
    ws['E3'] = "As of: 0900 Local"
    
    ws['B4'] = "UNIT POC:"
    ws['C4'] = "Program Manager"
    
    # Headers row (row 9 in example)
    headers = [
        'LAST NAME', 'FIRST NAME', 'GRADE', 'EDIPI', 'ASSIGNED LOCATION',
        'STATUS', 'PHYSICAL LOCATION', 'SERVICE', 'COMPONENT', 
        'SECTION/UNIT', 'DEPLOY START DATE', 'DEPART DATE', 'Citizenship\n(US or OCN)'
    ]
    
    header_row = 9
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align
    
    # Data rows
    data_row = 10
    for p in personnel:
        last_name, first_name = split_name(p.get('full_name', ''))
        
        # Assigned location
        assigned_loc = p.get('country_name') or p.get('current_location') or 'N/A'
        
        # Physical location
        physical_loc = p.get('camp_name') or p.get('current_location') or ''
        
        # Status
        status = p.get('status', 'Present')
        if status == 'Active':
            status = 'Present'
        
        # Citizenship - default US, could be OCN for non-US
        citizenship = p.get('nationality', 'US')
        if citizenship and citizenship.upper() not in ['US', 'USA', 'AMERICAN']:
            citizenship = 'OCN'
        else:
            citizenship = 'US'
        
        # Deploy date
        deploy_date = p.get('bog_date') or ''
        
        row_data = [
            last_name,                          # A - Last Name
            first_name,                         # B - First Name
            'Contractor',                       # C - Grade
            'N/A',                              # D - EDIPI
            assigned_loc,                       # E - Assigned Location
            status,                             # F - Status
            physical_loc,                       # G - Physical Location
            'ARMY',                             # H - Service
            'USACE',                            # I - Component
            'HNC/TFS',                          # J - Section/Unit
            deploy_date,                        # K - Deploy Start Date
            '',                                 # L - Depart Date
            citizenship                         # M - Citizenship
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=data_row, column=col, value=value)
            cell.border = thin_border
            if col <= 2:
                cell.alignment = left_align
            else:
                cell.alignment = center_align
        
        data_row += 1
    
    # Adjust column widths
    col_widths = [18, 12, 10, 12, 12, 8, 25, 8, 8, 10, 15, 12, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Set row height for header
    ws.row_dimensions[header_row].height = 30
    
    wb.save(output_path)
    print(f"Excel report generated: {output_path}")


def main():
    parser = argparse.ArgumentParser(description='Generate PERSTAT reports')
    parser.add_argument('--format', choices=['pdf', 'excel'], required=True,
                       help='Output format (pdf or excel)')
    parser.add_argument('--output', required=True, help='Output file path')
    
    args = parser.parse_args()
    
    if args.format == 'pdf':
        generate_pdf_report(args.output)
    else:
        generate_excel_report(args.output)


if __name__ == '__main__':
    main()
